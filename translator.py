
import sys
import openpyxl
import xlsxwriter

# import wikipron
# from wikipron.scrape import _PAGE_TEMPLATE, HTTP_HEADERS
# import requests_html
# import requests

import eng_to_ipa as ipa

global word_dict
global words
global config

global map_file
global input_file
global output_file
global problem_words

def create_map(sh):
    temp = {}
    for row in sh:
        second = False
        for cell in row:
            if(cell.value == None):
                return temp
            if(second):
                c2 = cell.value
                temp[c1] = c2
            else:
                second = True
                c1 = cell.value
    return temp

def get_words(sh):
    temp = []
    for row in sh:
        for cell in row:
            if(cell.value == None):
                return temp
            temp.append(cell.value)
    
    return temp

# gets the ipa translation of a word from wiktionary
# user must be connected to the internet
def get_ipa_wikipron(word):
    html_session = requests_html.HTMLSession()
    response = html_session.get(
        _PAGE_TEMPLATE.format(word=word), headers=HTTP_HEADERS
    )

    v = config.extract_word_pron(word, response, config)
    for w, pron in v:
        return pron.replace(' ', '')

# gets the ipa translation of a word from the Carnegie-Mellon University Dictionary
def get_ipa_cmu(word):
    if not ipa.isin_cmu(word):
        return None
    return ipa.convert(word)

def produce_output(wrksht):
    issues = []

    row = 0
    column = 0

    wrksht.write(row, column, "word")
    column += 1
    wrksht.write(row, column, "ipa")
    for phenome in word_dict:
        column += 1
        wrksht.write(row, column, phenome + " (" + str(word_dict[phenome]) + ")")
    column += 1
    wrksht.write(row, column, "score")
    
    row = 0
    column = 0

    for word in words:
        ipa = get_ipa_cmu(word)
        if(ipa == None):
            issues.append(word)
            continue
    
        temp_column = column
        row += 1

        wrksht.write(row, temp_column, word)
        temp_column += 1
        wrksht.write(row, temp_column, ipa)
        score = 0
        for phenome in word_dict:
            temp_column += 1
            count = ipa.count(phenome)
            wrksht.write(row, temp_column, count)
            score += count * word_dict[phenome]
            ipa = ipa.replace(phenome, "")
        temp_column += 1
        wrksht.write(row, temp_column, score)
    
    return issues

if __name__ == "__main__":
    # if len(sys.argv) < 2:
    #     print('[ERROR] An ISO 639-3 language code must be provided (e.g., "eng" for English).')
    #     sys.exit(1)  
    # try:
    #     lang = sys.argv[1]
    #     config = wikipron.Config(key=lang, skip_spaces_pron=False)  
    # except:
    #     print("[ERROR] Invalid ISO 639-3 language code: " + lang)
    #     sys.exit(1)

    if len(sys.argv) < 2:
        print('[ERROR] A file for phonetic mapping must be provided')
        sys.exit(1) 
    try:
        map_file = sys.argv[1]
        wrkbk = openpyxl.load_workbook(map_file)
        sh = wrkbk.active
        word_dict = create_map(sh.iter_rows(min_row=0, min_col=0, max_row=50000, max_col=2))
    except Exception as e:
        print('[ERROR] An error occured while trying to read file for phonettic mapping: ' + map_file + "\n")
        print(e)
        sys.exit(1)

    if len(sys.argv) < 3:
        print('[ERROR] A file for words must be provided')
        sys.exit(1)
    try:
        input_file = sys.argv[2]
        wrkbk = openpyxl.load_workbook(input_file)
        sh = wrkbk.active
        words = get_words(sh.iter_rows(min_row=0, min_col=0, max_row=50000, max_col=1))
    except Exception as e:
        print('[ERROR] An error occured while trying to read file with a list of words: ' + input_file + "\n")
        print(e)
        sys.exit(1)
    
    if len(sys.argv) < 4:
        print('[ERROR] An empty output file must be provided.')
        sys.exit(1)
    try:
        output_file = sys.argv[3]
        wrkbk = xlsxwriter.Workbook(output_file)
        wrksht = wrkbk.add_worksheet()
        problem_words = produce_output(wrksht)
        wrkbk.close()
    except Exception as e:
        print('[ERROR] An error occured while reading output file ' + output_file + "\n")
        print(e)
        sys.exit(1)
    
    print('Translation Successfully Completed.')
    print('-----------------------------------')
    if(len(problem_words) > 0):
        print('Words that could not be translated: ')
        for word in problem_words:
            print(word)
